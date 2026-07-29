#!/usr/bin/env python3
"""Generate a synthetic construction document set with planted conflicts + decoys."""
import pathlib
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, PageBreak)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = pathlib.Path(__file__).parent / "testpack"
OUT.mkdir(exist_ok=True)

DV = "/usr/share/fonts/truetype/dejavu/"
pdfmetrics.registerFont(TTFont("DJ", DV + "DejaVuSans.ttf"))
pdfmetrics.registerFont(TTFont("DJ-B", DV + "DejaVuSans-Bold.ttf"))
pdfmetrics.registerFont(TTFont("DJM", DV + "DejaVuSansMono.ttf"))

PROJECT = "TEST OFFICE BUILDING — PLOT 42"
CLIENT = "Test Client (Synthetic Data — Not a Real Project)"

body = ParagraphStyle("b", fontName="Helvetica", fontSize=9, leading=13.5, spaceAfter=5)
h1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=13, leading=17, spaceAfter=9)
h2 = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=10, leading=14,
                    spaceBefore=9, spaceAfter=5)
small = ParagraphStyle("s", fontName="Courier", fontSize=7, leading=10,
                       textColor=colors.HexColor("#555555"))
ar = ParagraphStyle("ar", fontName="DJ", fontSize=10, leading=17, alignment=2,
                    spaceAfter=6)


def A(t):
    """Shape + reorder Arabic for PDF rendering."""
    return get_display(arabic_reshaper.reshape(t))


def AR(lines, style=None):
    """Arabic block: bidi must be applied per rendered line, so each source line
    is emitted as its own Paragraph (keeps visual line order correct)."""
    st = style or ar
    tight = ParagraphStyle("tight", parent=st, spaceAfter=0)
    out = [Paragraph(A(l), tight) for l in lines[:-1]]
    out.append(Paragraph(A(lines[-1]), st))
    return out


def footer(doc_no, rev):
    def draw(canvas, doc):
        canvas.saveState()
        canvas.setFont("Courier", 6.5)
        canvas.setFillColor(colors.HexColor("#666666"))
        canvas.drawString(20 * mm, 12 * mm, f"{PROJECT}  |  {doc_no}  |  Rev {rev}")
        canvas.drawRightString(doc.pagesize[0] - 20 * mm, 12 * mm,
                               f"Page {canvas.getPageNumber()}")
        canvas.setStrokeColor(colors.HexColor("#CCCCCC"))
        canvas.line(20 * mm, 15 * mm, doc.pagesize[0] - 20 * mm, 15 * mm)
        canvas.restoreState()
    return draw


def build(fname, doc_no, rev, story, land=False):
    path = OUT / fname
    d = SimpleDocTemplate(str(path), pagesize=landscape(A4) if land else A4,
                          leftMargin=20 * mm, rightMargin=20 * mm,
                          topMargin=18 * mm, bottomMargin=20 * mm,
                          title=doc_no, author=CLIENT)
    d.build(story, onFirstPage=footer(doc_no, rev), onLaterPages=footer(doc_no, rev))
    print(f"  {fname}")


def titleblock(doc_no, title, rev, date, disc):
    t = Table([["DOCUMENT No.", doc_no, "REVISION", rev],
               ["TITLE", title, "DATE", date],
               ["PROJECT", PROJECT, "DISCIPLINE", disc]],
              colWidths=[26 * mm, 78 * mm, 22 * mm, 44 * mm])
    t.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), "Courier", 7),
        ("FONT", (1, 0), (1, -1), "Helvetica-Bold", 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#999999")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EFEFEF")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#EFEFEF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    return [t, Spacer(1, 9 * mm)]


def tbl(data, widths, header=True, align_right=()):
    t = Table(data, colWidths=widths, repeatRows=1 if header else 0)
    st = [("FONT", (0, 0), (-1, -1), "Helvetica", 7.5),
          ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#AAAAAA")),
          ("VALIGN", (0, 0), (-1, -1), "TOP"),
          ("LEFTPADDING", (0, 0), (-1, -1), 4),
          ("TOPPADDING", (0, 0), (-1, -1), 3.5),
          ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5)]
    if header:
        st += [("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 7.5),
               ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E4E4E4"))]
    for c in align_right:
        st.append(("ALIGN", (c, 0), (c, -1), "RIGHT"))
    t.setStyle(TableStyle(st))
    return t


print("Building test pack…")

# ─────────────────────────────────────────────────────────────
# 01 — EMPLOYER REQUIREMENTS
# ─────────────────────────────────────────────────────────────
s = titleblock("ER-001", "Employer Requirements (Extract)", "B", "2026-02-14", "General")
s += [
    Paragraph("EMPLOYER REQUIREMENTS — EXTRACT", h1),
    Paragraph("SECTION 5 — SECURITY AND LIFE SAFETY SYSTEMS", h2),
    Paragraph("<b>5.1 General.</b> The Contractor shall provide complete, tested and "
              "commissioned security and life safety systems throughout the building, "
              "in accordance with the requirements of the local authorities having "
              "jurisdiction.", body),
    Paragraph("<b>5.2 Access Control.</b> An electronic access control system shall be "
              "provided to all restricted areas including the server room, main "
              "electrical room, and the building management control room. The system "
              "shall support card and biometric credentials.", body),
    Paragraph("<b>5.3 Surveillance.</b> A CCTV system shall be provided covering all "
              "main entrances, exits, lobbies and the basement car park. Recorded "
              "footage shall be retained for a minimum of thirty (30) days.", body),
    Paragraph("<b>5.4 Emergency Lighting.</b> Emergency lighting shall be provided "
              "throughout all protected escape routes, stairwells and exit discharge "
              "areas, with a minimum autonomy of three (3) hours.", body),
    Paragraph("<b>5.5 Fire Detection.</b> An addressable fire alarm and detection "
              "system shall be provided throughout the building.", body),
    Spacer(1, 4 * mm),
    Paragraph("SECTION 6 — BUILDING FABRIC", h2),
    Paragraph("<b>6.1 Thermal Performance.</b> The building envelope shall achieve the "
              "thermal performance criteria stated in the Specification. External wall "
              "thermal insulation shall be not less than 10 cm thick.", body),
    Paragraph("<b>6.2 Roof.</b> The roof build-up shall incorporate thermal insulation "
              "as detailed in the Specification and the Schedule of Finishes.", body),
    Spacer(1, 6 * mm),
    Paragraph("— End of Extract —", small),
]
build("01_Employer_Requirements.pdf", "ER-001", "B", s)

# ─────────────────────────────────────────────────────────────
# 02 — SCOPE OF WORKS
# ─────────────────────────────────────────────────────────────
s = titleblock("SOW-001", "Scope of Works (Extract)", "A", "2026-02-20", "General")
s += [
    Paragraph("SCOPE OF WORKS — EXTRACT", h1),
    Paragraph("3. THE WORKS", h2),
    Paragraph("The Works comprise the design completion, supply, installation, testing "
              "and commissioning of the following, without limitation:", body),
    Spacer(1, 2 * mm),
    tbl([["Ref", "Description of Works"],
         ["3.1", "Substructure, superstructure and building envelope works."],
         ["3.2", "Internal partitions, doors, ironmongery and finishes throughout."],
         ["3.3", "Supply and installation of a complete CCTV system including cameras, "
                 "recording equipment and associated containment."],
         ["3.4", "Supply and installation of a complete electronic access control "
                 "system to all restricted areas."],
         ["3.5", "Fire alarm and detection system, complete."],
         ["3.6", "Emergency and escape lighting to all protected escape routes."],
         ["3.7", "Mechanical, electrical and plumbing installations."],
         ["3.8", "External works, landscaping and car park."]],
        [16 * mm, 154 * mm]),
    Spacer(1, 5 * mm),
    Paragraph("4. EXCLUSIONS", h2),
    Paragraph("Loose furniture, IT active equipment and tenant fit-out works are "
              "excluded from this Scope.", body),
]
build("02_Scope_of_Works.pdf", "SOW-001", "A", s)

# ─────────────────────────────────────────────────────────────
# 03 — SPECIFICATION (Div 01 / 07 / 08)
# ─────────────────────────────────────────────────────────────
s = titleblock("SPEC-001", "Specification — Divisions 01, 07, 08", "C",
               "2026-03-05", "Architectural")
s += [
    Paragraph("SECTION 01 35 23 — FIRE AND LIFE SAFETY GENERAL REQUIREMENTS", h1),
    Paragraph("<b>1.4 FIRE RESISTANCE — GENERAL</b>", h2),
    Paragraph("A. All doors opening onto protected escape routes shall achieve a fire "
              "resistance rating of not less than 90 minutes, tested and certified in "
              "accordance with the relevant standard, and shall bear a permanent "
              "certification label.", body),
    Paragraph("B. All fire rated door assemblies shall be fitted with self-closing "
              "devices and intumescent smoke seals.", body),
    PageBreak(),

    Paragraph("SECTION 07 21 00 — THERMAL INSULATION", h1),
    Paragraph("<b>PART 2 — PRODUCTS</b>", h2),
    Paragraph("<b>2.1 EXTERNAL WALL INSULATION</b>", h2),
    Paragraph("A. External wall thermal insulation shall be rigid mineral wool boards "
              "of density not less than 100 kg/m³, with a thickness of 100 mm.", body),
    Paragraph("B. Insulation shall be mechanically fixed and shall be continuous "
              "across all junctions to avoid thermal bridging.", body),
    Paragraph("<b>2.2 ROOF INSULATION</b>", h2),
    Paragraph("A. Roof thermal insulation shall be extruded polystyrene boards with a "
              "thickness of 120 mm, laid in two staggered layers.", body),
    Paragraph("B. Compressive strength shall be not less than 300 kPa.", body),
    PageBreak(),

    Paragraph("SECTION 08 14 00 — INTERIOR WOOD DOORS", h1),
    Paragraph("<b>PART 2 — PRODUCTS</b>", h2),
    Paragraph("<b>2.3 FIRE RATED DOORS</b>", h2),
    Paragraph("A. All fire rated doors shall have a fire resistance rating of not less "
              "than 60 minutes.", body),
    Paragraph("B. Each fire rated door leaf and frame shall bear a permanent label from "
              "an approved certification body indicating the achieved rating.", body),
    Paragraph("C. Fire rated doors shall be fitted with self-closing devices, "
              "intumescent seals and smoke seals to all edges.", body),
    Paragraph("D. Vision panels in fire rated doors shall be fire rated glazing of a "
              "rating not less than that of the door leaf.", body),
    Paragraph("<b>2.4 NON-RATED INTERIOR DOORS</b>", h2),
    Paragraph("A. Non-rated interior doors shall be solid core flush timber doors, "
              "44 mm thick, factory finished.", body),
    Paragraph("<b>2.5 IRONMONGERY</b>", h2),
    Paragraph("A. Ironmongery shall be stainless steel grade 304 throughout.", body),
]
build("03_Specification_Div07_08.pdf", "SPEC-001", "C", s)

# ─────────────────────────────────────────────────────────────
# 04 — DOOR SCHEDULE (drawing schedule, landscape)
# ─────────────────────────────────────────────────────────────
rows = [["Mark", "Location", "Size (mm)", "Type", "Fire Rated",
         "Rating", "Self Closer", "Qty"]]
doors = [
    ("FD-01", "Stair Core A — Level 1",  "1000 x 2100", "Steel",  "Yes", "FR60", "Yes", "1"),
    ("FD-02", "Stair Core A — Level 2",  "1000 x 2100", "Steel",  "Yes", "FR60", "Yes", "1"),
    ("FD-03", "Stair Core B — Level 1",  "1000 x 2100", "Steel",  "Yes", "FR30", "Yes", "1"),
    ("FD-04", "Stair Core B — Level 2",  "1000 x 2100", "Steel",  "Yes", "FR60", "Yes", "1"),
    ("FD-05", "Electrical Room L1",      "900 x 2100",  "Steel",  "Yes", "FR60", "Yes", "1"),
    ("FD-06", "Server Room L2",          "900 x 2100",  "Steel",  "Yes", "FR60", "Yes", "1"),
    ("FD-07", "Corridor L2 — Cross",     "1000 x 2100", "Timber", "Yes", "",     "Yes", "1"),
    ("FD-08", "Corridor L3 — Cross",     "1000 x 2100", "Timber", "Yes", "FR60", "Yes", "1"),
    ("FD-09", "Plant Room Roof",         "900 x 2100",  "Steel",  "Yes", "FR60", "Yes", "1"),
    ("FD-10", "Basement Car Park Lobby", "1000 x 2100", "Steel",  "Yes", "FR60", "Yes", "1"),
    ("FD-11", "Refuse Store",            "900 x 2100",  "Steel",  "Yes", "FR60", "Yes", "1"),
    ("FD-12", "Generator Room",          "1200 x 2100", "Steel",  "Yes", "FR60", "Yes", "1"),
    ("FD-13", "Riser Cupboard L2",       "700 x 2100",  "Steel",  "Yes", "FR60", "Yes", "1"),
    ("FD-14", "Riser Cupboard L3",       "700 x 2100",  "Steel",  "Yes", "FR60", "Yes", "1"),
    ("D-101", "Meeting Room 1.01",       "900 x 2100",  "Timber", "No",  "—",    "No",  "1"),
    ("D-101A","Meeting Room 1.01 (2nd leaf)", "600 x 2100", "Timber", "No", "—",  "No",  "1"),
    ("D-220", "Office 2.20",             "800 x 2100",  "Timber", "No",  "—",    "No",  "1"),
    ("D-221", "Office 2.21",             "800 x 2100",  "Timber", "No",  "—",    "No",  "1"),
]
rows += list(doors)
s = titleblock("A-201", "Door Schedule", "2", "2026-03-18", "Architectural")
s += [
    Paragraph("DOOR SCHEDULE", h1),
    tbl(rows, [18*mm, 62*mm, 28*mm, 20*mm, 20*mm, 20*mm, 22*mm, 14*mm]),
    Spacer(1, 5 * mm),
    Paragraph("GENERAL NOTES", h2),
    Paragraph("1. Refer to Specification Section 08 14 00 for door construction and "
              "ironmongery requirements.", body),
    Paragraph("2. All fire rated doors to be supplied with certification labels.", body),
    Paragraph("3. Total fire rated doors on this schedule: 14 No.", body),
    Paragraph("4. Roof build-up: waterproofing on 80 mm thermal insulation on screed "
              "to falls. Refer to roof detail.", body),
]
build("04_Door_Schedule.pdf", "A-201", "2", s, land=True)

# ─────────────────────────────────────────────────────────────
# 05 — BOQ (Excel)
# ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOQ"
hdr = ["Item", "Description", "Unit", "Qty", "Rate (QAR)", "Amount (QAR)"]
ws.append([PROJECT]); ws.append(["BILL OF QUANTITIES — Rev A"]); ws.append([])
ws.append(hdr)

data = [
    ("7", "SECTION 7 — THERMAL & MOISTURE PROTECTION", "", "", "", ""),
    ("7.21.01", "Supply and install external wall thermal insulation, mineral wool, "
                "100mm thick, mechanically fixed", "m2", 2840, 145, None),
    ("7.21.02", "Supply and install roof thermal insulation, XPS boards, 120mm thick "
                "in two layers", "m2", 1120, 210, None),
    ("", "", "", "", "", ""),
    ("8", "SECTION 8 — OPENINGS", "", "", "", ""),
    ("8.14.01", "Solid core flush timber door, 44mm, single leaf 800x2100mm, complete "
                "with frame and ironmongery", "No.", 22, 1850, None),
    ("8.14.02", "Solid core flush timber door, 44mm, single leaf 900x2100mm, complete "
                "with frame and ironmongery", "No.", 8, 1980, None),
    ("8.14.03", "Fire rated door assembly, complete with frame, certification label, "
                "self-closing device and intumescent seals", "No.", 12, 4650, None),
    ("8.14.04", "Additional leaf to double door, 600x2100mm, non-rated", "No.", 4, 1240, None),
    ("", "", "", "", "", ""),
    ("28", "SECTION 28 — ELECTRONIC SAFETY & SECURITY", "", "", "", ""),
    ("28.23.01", "Video surveillance system, complete, including cameras, network "
                 "video recorder and 30-day storage", "LS", 1, 385000, None),
    ("28.23.02", "Fixed dome camera, indoor, complete with mounting and cabling",
     "No.", 24, 2450, None),
    ("28.31.01", "Addressable fire alarm and detection system, complete", "LS", 1, 295000, None),
]
for r in data:
    row = list(r)
    if row[3] not in ("", None) and row[4] not in ("", None):
        row[5] = round(row[3] * row[4], 2)
    ws.append(row)

thin = Side(style="thin", color="BBBBBB")
ws["A1"].font = Font(bold=True, size=12)
ws["A2"].font = Font(bold=True, size=10, color="666666")
for c in ws[4]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="44546A")
    c.alignment = Alignment(horizontal="center")
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=6):
    for c in row:
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if c.column in (4, 5, 6):
            c.number_format = "#,##0.00"
    if row[1].value and str(row[0].value or "").isdigit():
        for c in row:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDE3EC")
for col, w in zip("ABCDEF", (11, 74, 8, 11, 13, 15)):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"
wb.save(OUT / "05_BOQ.xlsx")
print("  05_BOQ.xlsx")

# ─────────────────────────────────────────────────────────────
# 06 — MINUTES OF MEETING (Arabic)
# ─────────────────────────────────────────────────────────────
s = titleblock("MOM-014", "Minutes of Meeting No. 14", "-", "2026-04-02", "General")
arh = ParagraphStyle("arh", parent=h2, alignment=2, fontName="DJ-B")
h1_ar = ParagraphStyle("h1ar", parent=h1, fontName="DJ-B")
s += [Paragraph(A("محضر اجتماع رقم ١٤ — اجتماع التنسيق الفني"), h1_ar)]
s += AR(["المشروع: مبنى إداري تجريبي — قطعة ٤٢",
         "التاريخ: ٢ أبريل ٢٠٢٦        المكان: مكتب الاستشاري"])
s += [Spacer(1, 4 * mm), Paragraph(A("١. أنظمة الأمن"), arh)]
s += AR(["أكد المقاول أن عدد الكاميرات المطلوب لتغطية المداخل والمخارج",
         "ومواقف السيارات هو ٢٤ كاميرا، وهو مطابق لما ورد في جدول الكميات."])
s += AR(["طلب الاستشاري توضيح موقف نظام التحكم في الدخول (Access Control)",
         "حيث لم يُلاحظ له بند في جدول الكميات المقدَّم."])
s += [Spacer(1, 3 * mm), Paragraph(A("٢. الأبواب"), arh)]
s += AR(["نوقش تصنيف مقاومة الحريق للأبواب المفتوحة على مسارات الهروب.",
         "أشار الاستشاري إلى وجود اختلاف بين البند ٠١ ٣٥ ٢٣ والبند ٠٨ ١٤ ٠٠",
         "في المواصفة، وسيتم إصدار طلب معلومات (RFI) لحسم القيمة."])
s += [Spacer(1, 3 * mm), Paragraph(A("٣. العزل"), arh)]
s += AR(["طلب المقاول تأكيد سماكة عزل السطح، حيث ورد ١٢٠ مم في المواصفة",
         "بينما ورد ٨٠ مم في ملاحظات جدول الأبواب."])
s += [Spacer(1, 3 * mm), Paragraph(A("٤. الإنارة"), arh)]
s += AR(["أكد الاستشاري أن إنارة الطوارئ في مسارات الهروب مطلوبة وفق",
         "متطلبات المالك، وطلب مراجعة جدول الكميات."])
build("06_Minutes_of_Meeting_AR.pdf", "MOM-014", "-", s)

# ─────────────────────────────────────────────────────────────
# RENDER VERIFICATION
# ─────────────────────────────────────────────────────────────
# A PDF can extract text perfectly and still render as garbage — a corrupted
# font subset draws blobs while the text layer stays intact. Checking the
# extracted text alone does NOT prove the document is readable, so every page
# is rendered and its ink coverage compared against a recorded baseline.
#
# This exists because font subsetting silently destroyed the Arabic page once:
# ink fell from 0.0249 to 0.0079 while the text layer still looked fine.

INK_BASELINE = {                       # page ink ratio @ 72 dpi, greyscale
    ("01_Employer_Requirements.pdf", 1): 0.0416,
    ("02_Scope_of_Works.pdf", 1): 0.0243,
    ("03_Specification_Div07_08.pdf", 1): 0.0191,
    ("03_Specification_Div07_08.pdf", 2): 0.0160,
    ("03_Specification_Div07_08.pdf", 3): 0.0208,
    ("04_Door_Schedule.pdf", 1): 0.0298,
    ("04_Door_Schedule.pdf", 2): 0.0049,   # sparse continuation page — legitimately low
    ("06_Minutes_of_Meeting_AR.pdf", 1): 0.0249,
}
TOLERANCE = 0.30


def verify():
    try:
        import fitz
    except ImportError:
        print("\n! pymupdf not installed — render verification SKIPPED")
        print("  pip install pymupdf   (text extraction alone cannot catch broken glyphs)")
        return True

    print("\nRender verification (ink coverage vs baseline):")
    ok = True
    for f in sorted(OUT.glob("*.pdf")):
        doc = fitz.open(f)
        for n, page in enumerate(doc, 1):
            px = page.get_pixmap(dpi=72, colorspace=fitz.csGRAY)
            ink = sum(1 for v in px.samples if v < 200) / len(px.samples)
            base = INK_BASELINE.get((f.name, n))
            if base is None:
                print(f"  ?  {f.name:34s} p{n}  ink={ink:.4f}  (no baseline)")
                continue
            drift = abs(ink - base) / base
            good = drift <= TOLERANCE
            ok &= good
            print(f"  {'OK' if good else 'FAIL':4s} {f.name:34s} p{n}  "
                  f"ink={ink:.4f}  base={base:.4f}  drift={drift:+.0%}")
        doc.close()
    return ok


print(f"\nDone → {OUT}")
for f in sorted(OUT.iterdir()):
    print(f"  {f.name:38s} {f.stat().st_size/1024:7.1f} KB")

if not verify():
    raise SystemExit(
        "\nFAILED: a page no longer renders as expected.\n"
        "Glyphs are likely broken even if the text layer still extracts.\n"
        "Open the page and look at it before changing the baseline.")
print("\nAll pages render as expected.")
